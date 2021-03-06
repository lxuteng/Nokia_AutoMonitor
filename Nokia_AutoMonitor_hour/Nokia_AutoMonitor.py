import pandas
from pandas.core.frame import DataFrame
import openpyxl
import os
import sys
import multiprocessing
import cx_Oracle
import traceback
import datetime
import sqlite3
import pyDes
# import random
import csv
import time
from io import BytesIO
import base64
from matplotlib import pyplot as plt

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

font = {'family': 'SimHei',
        'weight': 'bold',
        # 'size': '16'
        }
plt.rc('font', **font)
plt.rc('axes', unicode_minus=False)


def copy_right():
    print('\n')
    print(u"""
    ------------------------------------------
    -   Welcome to use tools!                -
    -   Author : xuteng.lin                  -
    -   E_mail : xuteng.lin@huanuo-nsb.com   -
    ------------------------------------------
    """)
    print('\n')
    auth_time = int(time.strftime('%Y%m%d', time.localtime(time.time())))

    self_key = 'lxtnokia'
    self_iv = 'nokialxt'
    main_path = os.path.split(os.path.abspath(sys.argv[0]))[0]
    try:
        temp_license = open(os.path.join(main_path, 'license')).read()
    except:
        print(u'>>> 无license文件，请联系作者申请!')
        os.system("pause")
        sys.exit()
    k = pyDes.des(self_key,
                  mode=pyDes.CBC,
                  IV=self_iv,
                  pad=None,
                  padmode=pyDes.PAD_PKCS5)
    decryptstr = str(k.decrypt(base64.b64decode(temp_license)),
                     encoding='utf-8').split('-')
    if decryptstr[3] == 'Nokia_AutoMonitor_hour':
        if auth_time > int(decryptstr[2]):
            print(u'>>> 试用版本已过期，请更新！')
            os.system("pause")
            sys.exit()
    else:
        print(u'>>> license错误，请联系作者申请！')
        os.system("pause")
        sys.exit()

    print(u'''
    update log:
    2019-01-26 初版；
    2019-01-28 修补已知bug；
    2019-01-28 邮件正文添加TOP小区个数通报；显示细节调整；
    2019-03-02 优化TOP小区保存机制；
    ''')
    print(u'-' * 36)
    print(u'      >>>   starting   <<<')
    print(u'-' * 36)
    print(u'\n')
    time.sleep(1)


class Main:

    def __init__(self):
        """初始化"""
        copy_right()
        self.main_path = os.path.split(os.path.abspath(sys.argv[0]))[0]
        # 获取配置文件
        self.config_list = {}
        self.get_config()
        self.temp_time_now = datetime.datetime.now()
        self.temp_time = self.temp_time_now.strftime('%Y_%m_%d_%H_%M_%S')
        self.add_type_init = 0

        # 统计top小区个数
        self.num_top_cell_list = {}

    def get_config(self):
        # 初始化配置列表

        path_base_data = os.path.join(
            self.main_path,
            '_config',
            'config.xlsx')
        f_base_data_wb = openpyxl.load_workbook(path_base_data, read_only=True)
        for temp_sheet_name in f_base_data_wb.sheetnames:
            if temp_sheet_name not in self.config_list:
                self.config_list[temp_sheet_name] = {}
                temp_f_base_data_wb_sheet = f_base_data_wb[temp_sheet_name]
                if temp_sheet_name == 'sql_db':
                    temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
                    next(temp_iter_rows)
                    for temp_row in temp_iter_rows:
                        temp_value = [j.value for j in temp_row]
                        if temp_value[4] == '启用':
                            self.config_list[temp_sheet_name][temp_value[0]] = [
                                temp_value[1],
                                temp_value[2],
                                temp_value[3]
                            ]
                elif temp_sheet_name == 'sql_script_map':
                    temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
                    next(temp_iter_rows)
                    for temp_row in temp_iter_rows:
                        temp_value = [j.value for j in temp_row]
                        if temp_value[2] == '启用':
                            self.config_list[temp_sheet_name][temp_value[
                                1]] = [temp_value[0],
                                       temp_value[2],
                                       temp_value[3],
                                       temp_value[4],
                                       ]

                elif temp_sheet_name == 'sql_script_map_local':
                    temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
                    next(temp_iter_rows)
                    for temp_row in temp_iter_rows:
                        temp_value = [j.value for j in temp_row]
                        if temp_value[2] == '启用':
                            self.config_list[temp_sheet_name][temp_value[
                                1]] = [temp_value[0],
                                       temp_value[3],
                                       temp_value[4],
                                       temp_value[5],
                                       temp_value[6],
                                       ]

                elif temp_sheet_name == 'report_table_range':
                    temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
                    next(temp_iter_rows)
                    for temp_row in temp_iter_rows:
                        temp_value = [j.value for j in temp_row]
                        self.config_list[temp_sheet_name][temp_value[
                            0]] = temp_value[1:]

                elif temp_sheet_name == '生成趋势图kpi':
                    self.config_list[temp_sheet_name] = []
                    temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
                    next(temp_iter_rows)
                    for temp_row in temp_iter_rows:
                        temp_value = [j.value for j in temp_row]
                        if temp_value[1] == '启用':
                            self.config_list[temp_sheet_name].append(
                                temp_value[0])

                else:
                    for temp_row in temp_f_base_data_wb_sheet.iter_rows():
                        temp_value = [j.value for j in temp_row]
                        self.config_list[
                            temp_sheet_name
                        ][temp_value[0]] = temp_value[1]

        if self.config_list['main']['本地数据库'] is None:
            self.config_list['main']['本地数据库'] = os.path.join(
                self.main_path,
                '_db',
            )

    def online_db_process(self):
        # 多进程到不同的数据库上获取数据
        try:
            if self.config_list['config']['多进程'] == 'max':
                processes_num = len(self.config_list['sql_db'])
            else:
                processes_num = self.config_list['config']['多进程']
            process_pool = multiprocessing.Pool(processes=processes_num)
            for temp_db in self.config_list['sql_db']:
                process_pool.apply_async(
                    self.online_db_operator,
                    args=(temp_db,
                          self.config_list['sql_db'][temp_db]
                          ),
                    callback=self.online_db_input_warehousing
                )
            process_pool.close()
            process_pool.join()
        except:
            traceback.print_exc()

    def online_db_operator(self, db_name, db_information):
        data_list = {'db_name': db_name}
        connect_state = 0
        cc = ''
        try:
            ip = db_information[0]
            user = db_information[1]
            pwd = db_information[2]
            print('>>> 尝试连接数据库 {0}:{1} ...'.format(db_name, ip))
            try:
                conn = cx_Oracle.connect(user, pwd, ip)
                cc = conn.cursor()
                connect_state = 1
                print('>>>>>> {0}:{1} 连接成功!'.format(db_name, ip))
            except:
                print('||| 无法连接数据库 {0}:{1},请检查网络或数据库设置!'.format(
                    db_name,
                    ip
                ))
        except:
            traceback.print_exc()

        if connect_state == 1:
            for temp_sql in self.config_list['sql_script_map']:
                print(db_name, ':正在获取online数据...', temp_sql)
                self.online_get_data(
                    self.get_sql_text(
                        temp_sql,
                        'online',
                        self.config_list['sql_script_map'][temp_sql][0],
                        db_name
                    ), cc)

                try:
                    temp_data = DataFrame(cc.fetchall())
                    if temp_data is not None:
                        temp_head = [i[0] for i in cc.description]
                        temp_data.rename(
                            columns={
                                temp_head.index(i): i for i in temp_head
                            },
                            inplace=True
                        )
                        if len(temp_data) != 0:
                            data_list[temp_sql] = temp_data
                            # 保存采集记录
                            if self.config_list['sql_script_map'][temp_sql][
                                    0] not in ['para', ]:

                                # csv版本
                                with open(
                                        path_base_data_log, 'a', newline=''
                                ) as f:
                                    f_csv_w = csv.writer(f)
                                    for temp_sdate in sorted(list(set(
                                            temp_data.SDATE))):
                                        temp_low = [
                                            db_name,
                                            temp_sql,
                                            temp_sdate,
                                        ]
                                        f_csv_w.writerow(temp_low)

                                # excel版本
                                # f_base_data_wb = openpyxl.load_workbook(
                                #     path_base_data_log,
                                #     read_only=False)
                                #
                                # temp_f_base_data_wb_sheet = f_base_data_wb[
                                #     '数据采集记录表']
                                # for temp_sdate in sorted(list(set(
                                #         temp_data.SDATE))):
                                #     temp_low = [
                                #         db_name,
                                #         temp_sql,
                                #         temp_sdate,
                                #     ]
                                #     temp_f_base_data_wb_sheet.append(temp_low)
                                # f_base_data_wb.save(path_base_data_log)

                except:
                    pass
                    # traceback.print_exc()

        print(db_name, ':数据获取完毕，待入库...')
        return data_list

    def get_sql_text(self, sql_name, online_type, time_type, temp_cp):
        sql_text = ''
        temp_sql_full_name = os.path.join(
            self.main_path,
            '_sql',
            online_type,
            ''.join((sql_name, '.sql'))
        )
        try:
            f = open(temp_sql_full_name)
            temp_time_list = self.get_sql_between_time(time_type, temp_cp,
                                                       sql_name)
            temp_start_time = temp_time_list[0]
            temp_end_time = temp_time_list[1]
            temp_time_item = temp_time_list[2]
            temp_time_item_format = ''
            for temp_item in temp_time_item:
                temp_time_item_format += r"to_date({0}, 'yyyymmddhh24')," \
                                         r"".format(temp_item)

            sql_text = f.read().replace('&1', temp_start_time).replace(
                '&2', temp_end_time).replace('&3', temp_time_item_format[:-1])
        except:
            traceback.print_exc()
        return sql_text

    @staticmethod
    def online_get_data(sql_text, cc):
        try:
            return cc.execute(sql_text)
        except:
            traceback.print_exc()

    def get_sql_between_time_simp(self, time_type, int_num):
        start_time = 0
        end_time = 0
        if time_type == 'hour':
            start_time = (self.temp_time_now - datetime.timedelta(
                hours=int_num)).strftime('%Y%m%d%H')
            end_time = self.temp_time_now.strftime('%Y%m%d%H')

        return str(start_time), str(end_time)

    def get_sql_between_time_item(self, time_num):
        time_item_list = []
        for temp_i in range(time_num + 1):
            temp_time_item = (self.temp_time_now - datetime.timedelta(
                hours=temp_i)).strftime('%Y%m%d%H')
            time_item_list.append(temp_time_item)

        return time_item_list

    def get_sql_between_time(self, time_type, temp_cp, temp_sql_name):
        start_time = 0
        end_time = 0
        time_item = []
        exit_time_list = {}
        global path_base_data_log

        # csv版本
        path_base_data_log = os.path.join(
            self.main_path,
            '_db',
            '数据采集记录表.csv')
        try:
            with open(path_base_data_log, 'r') as f:
                f_csv = csv.reader(f)
                for temp_value in f_csv:
                    if temp_value[0] not in exit_time_list:
                        exit_time_list[temp_value[0]] = {}
                    if temp_value[1] not in exit_time_list[temp_value[0]]:
                        exit_time_list[temp_value[0]][temp_value[1]] = {}
                    exit_time_list[
                        temp_value[0]][temp_value[1]][temp_value[2]] = 0
        except:
            pass
            # traceback.print_exc()

        # excel版本
        # path_base_data_log = os.path.join(
        #     self.main_path,
        #     '_db',
        #     '数据采集记录表.xlsx')
        # f_base_data_wb = openpyxl.load_workbook(
        #     path_base_data_log, read_only=True)
        # for temp_sheet_name in f_base_data_wb.sheetnames:
        #     if temp_sheet_name not in exit_time_list:
        #         exit_time_list[temp_sheet_name] = {}
        #     temp_f_base_data_wb_sheet = f_base_data_wb[temp_sheet_name]
        #     if temp_sheet_name == '数据采集记录表':
        #         temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
        #         next(temp_iter_rows)
        #         for temp_row in temp_iter_rows:
        #             temp_value = [j.value for j in temp_row]
        #             if temp_value[0] not in exit_time_list[temp_sheet_name]:
        #                 exit_time_list[temp_sheet_name][temp_value[0]] = {}
        #             if temp_value[1] not in exit_time_list[temp_sheet_name][
        #                 temp_value[0]]:
        #                 exit_time_list[temp_sheet_name][temp_value[
        #                     0]][temp_value[1]] = {}
        #             exit_time_list[temp_sheet_name][temp_value[
        #                 0]][temp_value[1]][temp_value[2]] = temp_value[3]

        if time_type == 'hour':
            start_time = (self.temp_time_now - datetime.timedelta(
                hours=self.config_list['main']['获取最近连续时段数']
            )).strftime('%Y%m%d%H')
            end_time = self.temp_time_now.strftime('%Y%m%d%H')

            # 获取独立时段
            for temp_i in range(self.config_list['main']['获取最近连续时段数']+1):
                temp_time_item = (self.temp_time_now - datetime.timedelta(
                    hours=temp_i)).strftime('%Y%m%d%H')
                try:
                    if temp_time_item not in exit_time_list[
                            temp_cp][temp_sql_name]:
                        time_item.append(temp_time_item)
                except:
                    time_item.append(temp_time_item)

        return str(start_time), str(end_time), set(time_item)

    def online_db_input_warehousing(self, temp_data_list):
        db_name = temp_data_list['db_name']
        try:
            local_conn = sqlite3.connect(
                os.path.join(self.main_path, '_db', 'db.db'),
                check_same_thread=False
            )
            for temp_sql in temp_data_list:
                if temp_sql != 'db_name':
                    print(db_name, ':入库中...', temp_sql)
                    add_type = 'append'
                    if self.config_list['sql_script_map'][temp_sql][2] == '覆盖':
                        if self.add_type_init == 0:
                            add_type = 'replace'
                    temp_data_list[temp_sql].to_sql(
                        temp_sql,
                        con=local_conn,
                        if_exists=add_type,
                        chunksize=500
                    )
                    print(db_name, ':入库完毕...', temp_sql)

            local_conn.close()
            self.add_type_init = 1
        except:
            traceback.print_exc()

    def write_kpi_detail(self, report_name, temp_cu):
        detail_list = 0
        if self.config_list[
                'sql_script_map_local'][report_name][4] == '累计保存':
            detail_list = 1

        try:
            kpi_list = os.path.join(
                self.main_path,
                '_report',
                ''.join(
                    (
                        'Detail_list_',
                        report_name,
                        '.csv'
                     )
                )
            )

            global kpi_list_temp
            kpi_list_temp = os.path.join(
                self.main_path,
                '_report',
                ''.join((
                    'All_kpi_detail_list_',
                    self.temp_time,
                    '.xlsx'
                ))
            )
            # if not os.path.exists(kpi_list):
            #     workbook_a = openpyxl.Workbook()
            #     workbook_a.save(kpi_list)
            # workbook_a = openpyxl.load_workbook(kpi_list)

            if not os.path.exists(kpi_list_temp):
                workbook = openpyxl.Workbook()
                workbook.save(kpi_list_temp)
            workbook = openpyxl.load_workbook(kpi_list_temp)

            temp_head = [i[0] for i in temp_cu.description]
            if detail_list:
                if not os.path.exists(kpi_list):
                    f = open(kpi_list, 'a', newline='')
                    f_w = csv.writer(f)
                    f_w.writerow(temp_head)
                else:
                    f = open(kpi_list, 'a', newline='')
                    f_w = csv.writer(f)
            worksheet = workbook.create_sheet(report_name)
            worksheet.append(temp_head)

            temp_value_list = temp_cu.fetchall()
            for temp_row in temp_value_list:
                if detail_list:
                    f_w.writerow(temp_row)
                worksheet.append(temp_row)

            workbook.save(kpi_list_temp)

            if self.config_list[
                    'sql_script_map_local'][report_name][3] == '呈现':
                self.get_top_cell_num(temp_value_list, report_name)

        except:
            traceback.print_exc()

    def local_db_operator(self):
        local_conn = sqlite3.connect(
            os.path.join(self.main_path, '_db', 'db.db'),
            check_same_thread=False
        )

        cu = local_conn.cursor()
        for temp_local_sql in self.config_list['sql_script_map_local']:
            print('='*32)
            print('>>>获取本地数据', temp_local_sql)
            f_sql = open(
                os.path.join(
                    self.main_path,
                    '_sql/local',
                    ''.join((temp_local_sql, '.sql'))
                ), encoding='utf-8-sig'
            )
            sql_scr = f_sql.read()

            if self.config_list[
                    'sql_script_map_local'][temp_local_sql][1] == '需要':
                temp_time_list = self.get_sql_between_time_simp(
                    self.config_list['sql_script_map_local'][temp_local_sql][0],
                    self.config_list['sql_script_map_local'][temp_local_sql][2],
                )
                temp_start_time = temp_time_list[0]
                temp_end_time = temp_time_list[1]
                sql_scr = sql_scr.replace(
                    '&1', temp_start_time).replace('&2', temp_end_time)
            try:
                cu.execute(sql_scr)
                self.write_kpi_detail(temp_local_sql, cu)
            except:
                traceback.print_exc()

    def report_table_class(self, kpi_name, kpi_value):
        temp_table_str = """      <td>"""
        temp_table_str += str(kpi_value)
        temp_table_str += """</td>\n"""

        if kpi_name in self.config_list['report_table_range']:

            if self.config_list['report_table_range'][kpi_name][0] == '表头':
                temp_table_str = """      <th>"""
                temp_table_str += str(kpi_value)
                temp_table_str += """</th>\n"""

            elif self.config_list['report_table_range'][kpi_name][0] == '<':
                if self.config_list['report_table_range'][kpi_name][2] is None:
                    if kpi_value <= self.config_list[
                            'report_table_range'][kpi_name][1]:
                        temp_table_str = """      <td bgcolor="#ffee93">"""
                        temp_table_str += """<b><font color="#B8860B">"""
                        temp_table_str += str(kpi_value)
                        temp_table_str += """</font></b></td>\n"""
                else:
                    if kpi_value <= self.config_list['report_table_range'][
                            kpi_name][2]:
                        temp_table_str = """      <td bgcolor="#ffc09f"><b>"""
                        temp_table_str += """<font color="#FF0000">"""
                        temp_table_str += str(kpi_value)
                        temp_table_str += """</font></b></td>\n"""
                    elif kpi_value <= self.config_list['report_table_range'][
                            kpi_name][1]:
                        temp_table_str = """      <td bgcolor="#ffee93"><b>"""
                        temp_table_str += """<font color="#B8860B">"""
                        temp_table_str += str(kpi_value)
                        temp_table_str += """</font></b></td>\n"""

            elif self.config_list['report_table_range'][kpi_name][0] == '>':
                if self.config_list['report_table_range'][kpi_name][2] is None:
                    if kpi_value >= self.config_list[
                            'report_table_range'][kpi_name][1]:
                        temp_table_str = """      <td bgcolor="#ffee93"><b>"""
                        temp_table_str += """<font color="#B8860B">"""
                        temp_table_str += str(kpi_value)
                        temp_table_str += """</font></b></td>\n"""
                else:
                    if kpi_value >= self.config_list['report_table_range'][
                            kpi_name][2]:
                        temp_table_str = """      <td bgcolor="#ffc09f"><b>"""
                        temp_table_str += """<font color="#FF0000">"""
                        temp_table_str += str(kpi_value)
                        temp_table_str += """</font></b></td>\n"""
                    elif kpi_value >= self.config_list['report_table_range'][
                            kpi_name][1]:
                        temp_table_str = """      <td bgcolor="#ffee93"><b>"""
                        temp_table_str += """<font color="#B8860B">"""
                        temp_table_str += str(kpi_value)
                        temp_table_str += """</font></b></td>\n"""
        return temp_table_str

    def to_html(self, df):

        table_str = """<table border="1" bordercolor="#B0C4DE" """
        table_str += """class="dataframe" cellspacing="0" """
        table_str += """style="text-align: "center">\n"""
        table_str += """  <thead>\n"""

        temp_columns_name_list = df.columns.values.tolist()

        table_str += """    <tr style="background-color:#EEF0F4"">\n"""

        for temp_head_row in temp_columns_name_list:
            table_str += """      <th>"""
            table_str += str(temp_head_row)
            table_str += """</th>\n"""

        table_str += """    </tr>\n"""
        table_str += """  </thead>\n"""
        table_str += """  <tbody style="background-color:#F6F8FA">\n"""

        for temp_index, temp_row in df.iterrows():
            table_str += """    <tr>\n"""
            temp_value_dict = dict(temp_row)
            for temp_columns_name in temp_columns_name_list:
                table_str += self.report_table_class(
                    temp_columns_name,
                    temp_value_dict[temp_columns_name]
                )
            table_str += """    </tr>\n"""

        table_str += """  </tbody>\n"""
        table_str += """</table>\n"""

        return table_str

    def to_image(self, df, kpi_name):
        # 设置画图长宽
        plt.rcParams['figure.figsize'] = (6.0, 3.0)
        # 清除前面图表
        plt.figure()

        temp_x, temp_data_list = self.to_image_df(df, kpi_name)
        for temp_data in temp_data_list:
            plt.plot(temp_x, temp_data[0], label=temp_data[1])

        # plt.ylim(98,100)
        plt.legend()
        # plt.legend(loc='upper left')
        plt.title(kpi_name)
        # plt.show()
        # plt.grid(True)
        buffer = BytesIO()
        plt.savefig(buffer)
        plot_data = buffer.getvalue()
        imb = base64.b64encode(plot_data)
        ims = imb.decode()
        imd = "data:image/png;base64," + ims
        iris_im = """<img src="%s">""" % imd
        return iris_im

    @staticmethod
    def to_image_df(df, kpi_name):

        temp_v = [str(i)[-2:] for i in list(df[df['CITY'] == '潮州'][
                                                'SDATE'])]
        temp_city_list = sorted(list(set(df['CITY'])))

        temp_data_list = [
            [df[df[u'CITY'] == temp_city][
                 [kpi_name]], temp_city] for temp_city in temp_city_list
        ]

        return temp_v, temp_data_list

    def report(self):
        df = pandas.read_excel(
            kpi_list_temp,
            sheet_name='city_cell_hour_all',
            index_col=False
        )
        temp_html = ''
        # 邮件介绍
        temp_html += """<body><h2><pre>Hi~ 各位领导、同事：</pre></h2>"""
        temp_html += """<h3><pre>   以下是最近时段诺基亚五地市的指标概况，"""
        temp_html += """请各位针对恶化指标进行优化处理，谢谢；</pre></h3>\n"""
        temp_html += """<h3><pre>   附件为各类TOP """
        temp_html += """kpi恶化小区供优化参考，其中：</pre></h3>\n"""
        # 邮件正文添加top小区个数通报
        for temp_table_name in self.num_top_cell_list:
            temp_num = 0
            temp_html_top_cell_num = """<h4><pre>      """
            temp_html_top_cell_num += """<font color="#FF0000">"""
            temp_html_top_cell_num += temp_table_name
            temp_html_top_cell_num += """</font>"""
            temp_html_top_cell_num += "："
            for temp_city in sorted(list(self.num_top_cell_list[
                                             temp_table_name].keys())):
                if self.num_top_cell_list[temp_table_name][temp_city] != 0:
                    temp_num += 1
                    temp_html_top_cell_num += '{0}:{1}个；'.format(
                        temp_city,
                        self.num_top_cell_list[temp_table_name][temp_city]
                    )
            temp_html_top_cell_num += """</pre></h4>\n"""
            if temp_num != 0:
                temp_html += temp_html_top_cell_num

        temp_time_list = self.get_sql_between_time_simp(
            'hour', self.config_list['main']['报告报表呈现时段数'])
        temp_html += self.to_html(
            df[df.SDATE >= int(temp_time_list[0])]
        )
        for temp_kpi in self.config_list['生成趋势图kpi']:
            temp_html += self.to_image(df, temp_kpi)

        temp_html_name = os.path.join(
            self.main_path,
            '_report',
            ''.join((
                'All_kpi_detail_list_',
                self.temp_time,
                '.html'
            ))
        )
        with open(temp_html_name, 'w') as f_html:
            f_html.write(temp_html)

        return temp_html

    def email(self, in_mail_text):

        if self.config_list['e_mail']['启用Emali功能'] != '启用':
            return 'end_email'

        en_time = self.config_list['e_mail']['允许邮件发送时段'].split(',')
        if self.temp_time_now.strftime('%H') not in en_time:
            print('>>> 当前未在允许邮件发送时段内，邮件未发送！')
            return 'end_email'

        smtp_servers = self.config_list['e_mail']['smtp服务器']
        address_from = self.config_list['e_mail']['邮箱用户名']
        pass_word = self.config_list['e_mail']['邮箱密码']
        address_to = []
        address_cc = []
        if self.config_list['e_mail']['主送人员'] is not None:
            address_to = self.config_list['e_mail']['主送人员'].split(',')
        if self.config_list['e_mail']['抄送人员'] is not None:
            address_cc = self.config_list['e_mail']['抄送人员'].split(',')
        address_all = address_to + address_cc

        mail_subject = self.config_list['e_mail']['邮件主题']

        # 邮件主体
        text_part = MIMEText(in_mail_text, 'html', 'utf-8')

        # 邮件附件
        excel_file_name = os.path.split(kpi_list_temp)[-1]
        excel_part = MIMEApplication(open(kpi_list_temp, 'rb').read())
        excel_part.add_header('Content-Disposition', 'attachment',
                              filename=excel_file_name)

        # 生成邮件相关设置信息
        mail_message = MIMEMultipart()
        mail_message['Subject'] = mail_subject + '_' + self.temp_time

        mail_message['From'] = address_from

        mail_message['To'] = ";".join(address_to)

        if len(address_cc) != 0:
            mail_message['Cc'] = ";".join(address_cc)

        mail_message.attach(text_part)
        mail_message.attach(excel_part)

        mail_message["Accept-Language"] = "zh-CN"
        mail_message["Accept-Charset"] = "ISO-8859-1,utf-8"

        try:
            server = smtplib.SMTP(smtp_servers)

            # 解决 connection unexpectedly closed报错
            server.ehlo()
            server.starttls()

            server.login(address_from, pass_word)
            server.sendmail(address_from, address_all, mail_message.as_string())
            print('>>> 邮件发送完成！')
            server.quit()
        except smtplib.SMTPException as e:
            print('error:', e)

    def data_clean(self):
        if main.config_list['main']['是否自动清除数据'] == '是':
            print('正在清除数据历史数据...')
            for temp_table in self.config_list['sql_script_map']:
                if self.config_list['sql_script_map'][temp_table][1] == '启用':
                    if self.config_list['sql_script_map'][temp_table][
                            2] == '添加':
                        time_num = self.config_list[
                            'sql_script_map'][temp_table][3]
                        if time_num is not None:
                            time_item = self.get_sql_between_time_item(time_num)
                            time_item_text = ','.join(time_item)
                            f_sql = open(
                                os.path.join(
                                    self.main_path,
                                    '_sql/local',
                                    'data_clean.sql'),
                                encoding='utf-8-sig'
                            )
                            sql_scr = f_sql.read()
                            sql_scr = sql_scr.replace(
                                '&table',
                                temp_table).replace('&time_item',
                                                    time_item_text)
                            local_conn = sqlite3.connect(
                                os.path.join(self.main_path, '_db', 'db.db'),
                                check_same_thread=False
                            )
                            cu = local_conn.cursor()
                            cu.execute(sql_scr)
                            local_conn.commit()
                            cu.execute('VACUUM')
                            cu.close()
                            local_conn.close()
            print('清除历史数据完成！')

    def get_top_cell_num(self, temp_value, temp_table_name):
        if temp_table_name not in self.num_top_cell_list:
            self.num_top_cell_list[temp_table_name] = {}
        for temp_i in temp_value:
            if temp_i[1] in self.num_top_cell_list[temp_table_name]:
                self.num_top_cell_list[temp_table_name][temp_i[1]] += 1
            else:
                self.num_top_cell_list[temp_table_name][temp_i[1]] = 1


if __name__ == '__main__':
    multiprocessing.freeze_support()
    print(''.join((time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))))
    star_time = time.time()
    main = Main()
    main.online_db_process()
    main.local_db_operator()
    mail_text = main.report()
    main.email(mail_text)
    main.data_clean()
    print(''.join((time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))))
    print(''.join(('>>> 历时：', time.strftime(
        '%Y/%m/%d %H:%M:%S',
        time.gmtime(time.time() - star_time)
        )
    )))
